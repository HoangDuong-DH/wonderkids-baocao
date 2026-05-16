"""
Generate WonderKids assessment template XLSX.

Output: uploads/WonderKids-Template-DG-DauKy.xlsx

Format khớp với importer của WonderKids — Báo cáo đánh giá năng lực đầu kỳ.
Khi kéo-thả file này vào báo cáo HTML, parser sẽ tìm các anchor sau:
  • HỌ TÊN, NĂM SINH, LỚP, GIÁO VIÊN, NGÀY KIỂM TRA, NGÀY BÁO CÁO
  • ĐÁNH GIÁ THEO ĐƠN VỊ RIÊNG BIỆT (bảng đơn vị KTCB)
  • SO SÁNH KHẢ NĂNG GIẢI TOÁN CÓ/KHÔNG CÓ LỜI VĂN
  • ĐÁNH GIÁ THEO LĨNH VỰC (bảng năng lực tư duy)
  • SO SÁNH NĂNG LỰC GIẢI QUYẾT VẤN ĐỀ THEO CẤP ĐỘ
  • 20 câu trắc nghiệm (cột MỤC + ĐƠN VỊ + ĐÁP ÁN + KẾT QUẢ)
  • 10 câu tư duy (cột MỤC + ĐIỂM CHUẨN + TỔNG)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).resolve().parent.parent / "uploads" / "WonderKids-Template-DG-DauKy.xlsx"
OUT.parent.mkdir(parents=True, exist_ok=True)

wb = Workbook()
ws = wb.active
ws.title = "ĐÁNH GIÁ ĐẦU KỲ"

# ── styles ─────────────────────────────────────────────────────────────────
TITLE = Font(name="Calibri", size=14, bold=True, color="C00000")
H1    = Font(name="Calibri", size=12, bold=True, color="C00000")
H2    = Font(name="Calibri", size=11, bold=True, color="000000")
HEAD  = Font(name="Calibri", size=10, bold=True)
NORM  = Font(name="Calibri", size=10)
HINT  = Font(name="Calibri", size=9, italic=True, color="808080")

YEL  = PatternFill("solid", fgColor="FFF2CC")
ORG  = PatternFill("solid", fgColor="FCE4D6")
GRY  = PatternFill("solid", fgColor="EDEDED")
WHT  = PatternFill("solid", fgColor="FFFFFF")

thin = Side(border_style="thin", color="999999")
BOX  = Border(left=thin, right=thin, top=thin, bottom=thin)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ── column widths ──────────────────────────────────────────────────────────
widths = {"A": 8, "B": 16, "C": 12, "D": 10, "E": 4,
          "F": 12, "G": 30, "H": 10, "I": 12, "J": 8,
          "K": 4,  "L": 14, "M": 12, "N": 14, "O": 18}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# ── helpers ────────────────────────────────────────────────────────────────
def put(r, c, val, font=None, fill=None, align=None, border=None, span=None):
    cell = ws.cell(row=r, column=c, value=val)
    if font:   cell.font   = font
    if fill:   cell.fill   = fill
    if align:  cell.alignment = align
    if border: cell.border = border
    if span:
        rr, cc = span
        ws.merge_cells(start_row=r, start_column=c, end_row=r+rr-1, end_column=c+cc-1)
    return cell

def row_label(r, label, value="", *, fill=None):
    put(r, 1, label,  font=HEAD, fill=fill or GRY, align=LEFT,   border=BOX)
    put(r, 2, value,  font=NORM, fill=WHT,         align=LEFT,   border=BOX)

# ── R1: TITLE ──────────────────────────────────────────────────────────────
put(1, 1, "BÁO CÁO ĐÁNH GIÁ NĂNG LỰC ĐẦU KỲ — WONDERKIDS",
    font=TITLE, align=CENTER, span=(1, 10))
ws.row_dimensions[1].height = 26

# ── R3-R5: HEADER INFO ─────────────────────────────────────────────────────
row_label(3, "HỌ TÊN", "")
put(3, 4, "NĂM SINH", font=HEAD, fill=GRY, align=LEFT, border=BOX)
put(3, 5, "",         font=NORM, fill=WHT, align=LEFT, border=BOX)

row_label(4, "LỚP", "")
put(4, 4, "GIÁO VIÊN", font=HEAD, fill=GRY, align=LEFT, border=BOX)
put(4, 5, "",          font=NORM, fill=WHT, align=LEFT, border=BOX)

row_label(5, "NGÀY KIỂM TRA", "")
put(5, 4, "NGÀY BÁO CÁO", font=HEAD, fill=GRY, align=LEFT, border=BOX)
put(5, 5, "",             font=NORM, fill=WHT, align=LEFT, border=BOX)

# ── R7: SECTION 1 HEADING ──────────────────────────────────────────────────
put(7, 1, "KIẾN THỨC CƠ BẢN", font=H1, fill=YEL, align=LEFT, span=(1, 10))
ws.row_dimensions[7].height = 22

# ── R8-R9: Sub-headers row + column headers row ────────────────────────────
put(8, 1, "KẾT QUẢ ĐÁNH GIÁ THEO CÂU HỎI RIÊNG BIỆT",
    font=H2, fill=ORG, align=CENTER, span=(1, 4))
put(8, 6, "ĐÁNH GIÁ THEO ĐƠN VỊ RIÊNG BIỆT",
    font=H2, fill=ORG, align=CENTER, span=(1, 5))

for col, label in enumerate(["MỤC", "ĐƠN VỊ", "ĐÁP ÁN", "KẾT QUẢ"], start=1):
    put(9, col, label, font=HEAD, fill=GRY, align=CENTER, border=BOX)
for col, label in enumerate(
    ["ĐƠN VỊ", "PHÂN LOẠI", "TIÊU CHUẨN", "ĐIỂM THỰC TẾ", "TỶ LỆ"], start=6):
    put(9, col, label, font=HEAD, fill=GRY, align=CENTER, border=BOX)

# ── R10-R29: 20 câu trắc nghiệm + 5 dòng đơn vị + tổng ─────────────────────
# 20 câu (rows 10-29). Cột A=số câu, B=ĐƠN VỊ (A-E), C=ĐÁP ÁN, D=KẾT QUẢ (O/X)
# Đơn vị (cột B) prefill theo chuẩn CMS bài 20 câu lớp 1
cms_units = ["A", "C", "B", "E", "B", "B", "D", "A", "C", "B",
             "D", "D", "B", "E", "E", "A", "A", "D", "C", "A"]
for q in range(1, 21):
    r = 9 + q
    put(r, 1, q,                  font=NORM, align=CENTER, border=BOX)
    put(r, 2, cms_units[q-1],     font=NORM, align=CENTER, border=BOX)  # đơn vị A-E
    put(r, 3, "",                 font=NORM, align=CENTER, border=BOX)  # đáp án
    put(r, 4, "",                 font=NORM, align=CENTER, border=BOX)  # kết quả O/X

# 5 đơn vị (rows 10-14)
units = [
    ("A", "Các số đến 9",              5),
    ("B", "Phép cộng và phép trừ",     5),
    ("C", "Phân loại",                 3),
    ("D", "Đo lường",                  4),
    ("E", "Các kiến thức toán tư duy khác", 3),
]
for i, (code, name, std) in enumerate(units):
    r = 10 + i
    put(r, 6, code,  font=HEAD, fill=YEL, align=CENTER, border=BOX)
    put(r, 7, name,  font=NORM, align=LEFT,   border=BOX)
    put(r, 8, std,   font=NORM, align=CENTER, border=BOX)
    put(r, 9, "",    font=NORM, align=CENTER, border=BOX)
    put(r, 10, "",   font=NORM, align=CENTER, border=BOX)

# TỔNG (row 15)
put(15, 6, "TỔNG", font=HEAD, fill=GRY, align=CENTER, border=BOX, span=(1, 2))
put(15, 8, 20,     font=HEAD, fill=GRY, align=CENTER, border=BOX)
put(15, 9, "",     font=HEAD, fill=GRY, align=CENTER, border=BOX)
put(15, 10, "",    font=HEAD, fill=GRY, align=CENTER, border=BOX)

# ── R30-R33: BÀI TOÁN có lời / không lời ───────────────────────────────────
put(30, 6, "SO SÁNH KHẢ NĂNG GIẢI TOÁN CÓ/KHÔNG CÓ LỜI VĂN",
    font=H2, fill=ORG, align=CENTER, span=(1, 5))
for col, label in enumerate(
    ["PHÂN LOẠI", "TIÊU CHUẨN", "ĐIỂM THỰC TẾ", "TỶ LỆ"], start=6):
    put(31, col, label, font=HEAD, fill=GRY, align=CENTER, border=BOX)

bt_rows = [
    ("Bài toán có lời văn", 13),
    ("Bài toán cơ bản",     7),
]
for i, (name, std) in enumerate(bt_rows):
    r = 32 + i
    put(r, 6, name, font=NORM, align=LEFT,   border=BOX)
    put(r, 7, std,  font=NORM, align=CENTER, border=BOX)
    put(r, 8, "",   font=NORM, align=CENTER, border=BOX)
    put(r, 9, "",   font=NORM, align=CENTER, border=BOX)
put(34, 6, "TỔNG", font=HEAD, fill=GRY, align=CENTER, border=BOX)
put(34, 7, 20,     font=HEAD, fill=GRY, align=CENTER, border=BOX)
put(34, 8, "",     font=HEAD, fill=GRY, align=CENTER, border=BOX)
put(34, 9, "",     font=HEAD, fill=GRY, align=CENTER, border=BOX)

# ── R36: SECTION 2 HEADING ─────────────────────────────────────────────────
put(36, 1, "TƯ DUY TOÁN HỌC", font=H1, fill=YEL, align=LEFT, span=(1, 15))
ws.row_dimensions[36].height = 22

# ── R37: Sub-section headers ───────────────────────────────────────────────
put(37, 1,  "KẾT QUẢ ĐÁNH GIÁ THEO CÂU HỎI RIÊNG BIỆT",
    font=H2, fill=ORG, align=CENTER, span=(1, 4))
put(37, 6,  "ĐÁNH GIÁ THEO LĨNH VỰC",
    font=H2, fill=ORG, align=CENTER, span=(1, 6))
put(37, 13, "SO SÁNH NĂNG LỰC GIẢI QUYẾT VẤN ĐỀ THEO CẤP ĐỘ",
    font=H2, fill=ORG, align=CENTER, span=(1, 6))

# ── R38: column headers ────────────────────────────────────────────────────
for col, label in enumerate(["MỤC", "", "ĐIỂM CHUẨN", "TỔNG"], start=1):
    put(38, col, label, font=HEAD, fill=GRY, align=CENTER, border=BOX)
for col, label in enumerate(
    ["PHÂN LOẠI", "NỘI DUNG ĐÁNH GIÁ", "CÂU",
     "ĐIỂM CHUẨN", "ĐIỂM THỰC TẾ", "TỶ LỆ"], start=6):
    put(38, col, label, font=HEAD, fill=GRY, align=CENTER, border=BOX)
for col, label in enumerate(
    ["PHÂN LOẠI", "SỐ LƯỢNG CÂU", "CỤ THỂ",
     "ĐIỂM CHUẨN", "ĐIỂM THỰC TẾ", "TỶ LỆ"], start=13):
    put(38, col, label, font=HEAD, fill=GRY, align=CENTER, border=BOX)

# ── R39-R51: 10 câu tư duy (LEFT) + 6 năng lực (MID) + 3 cấp độ (RIGHT) ──
# 10 câu: rows 39-48. Cụm: Q1-Q7 cơ bản, Q8-Q9 nâng cao, Q10 sáng tạo.
# Cluster label nằm cùng dòng đầu của cụm (cột A).
q_data = [
    (1,  5, "NĂNG LỰC TƯ DUY CƠ BẢN"),
    (2,  5, ""),
    (3,  5, ""),
    (4,  5, ""),
    (5,  5, ""),
    (6,  5, ""),
    (7,  5, ""),
    (8,  7, "TƯ DUY NÂNG CAO"),
    (9,  7, ""),
    (10, 11, "Ý TƯỞNG SÁNG TẠO - GIẢI QUYẾT VẤN ĐỀ"),
]
for i, (q, std, cluster) in enumerate(q_data):
    r = 39 + i
    put(r, 1, cluster, font=HEAD if cluster else NORM,
        fill=YEL if cluster else WHT, align=LEFT, border=BOX)
    put(r, 2, q,   font=NORM, align=CENTER, border=BOX)
    put(r, 3, std, font=NORM, align=CENTER, border=BOX)
    put(r, 4, "",  font=NORM, align=CENTER, border=BOX)  # điểm đạt

# 6 năng lực (rows 39-44)
nl_data = [
    ("A. Giải quyết vấn đề",      "Áp dụng kiến thức để giải quyết bài toán mới",         "2,3,5,7,8,9",   34),
    ("B. Logic và quan sát",       "Tìm ra mối quan hệ, bản chất qua trực quan",           "1,3,4,6,10",    31),
    ("C. Phân tích và phân loại",  "Thu thập, phân loại thông tin để giải quyết vấn đề",   "1,2,5,6,7,8,9", 39),
    ("D. Số và ký hiệu",           "Vận dụng khái niệm, ký hiệu toán học",                 "4,5,6,7,10",    31),
    ("E. Không gian và hình học",  "Giải quyết bài toán hình học phẳng và không gian",     "2,3,7,9,10",    33),
    ("F. Suy luận toán học",       "Sử dụng quy nạp, diễn dịch",                           "4,5,8,9",       24),
]
for i, (name, desc, cau, std) in enumerate(nl_data):
    r = 39 + i
    put(r, 6,  name, font=NORM, align=LEFT,   border=BOX)
    put(r, 7,  desc, font=NORM, align=LEFT,   border=BOX)
    put(r, 8,  cau,  font=NORM, align=CENTER, border=BOX)
    put(r, 9,  std,  font=NORM, align=CENTER, border=BOX)
    put(r, 10, "",   font=NORM, align=CENTER, border=BOX)
    put(r, 11, "",   font=NORM, align=CENTER, border=BOX)

# TỔNG cho năng lực (row 45)
put(45, 6, "TỔNG", font=HEAD, fill=GRY, align=CENTER, border=BOX, span=(1, 3))
put(45, 9, 192,     font=HEAD, fill=GRY, align=CENTER, border=BOX)
put(45, 10, "",     font=HEAD, fill=GRY, align=CENTER, border=BOX)
put(45, 11, "",     font=HEAD, fill=GRY, align=CENTER, border=BOX)

# 3 cấp độ (rows 39-41) — cột M-R
cd_data = [
    ("KHÓ",        3, "8,9,10",  25),
    ("TRUNG BÌNH", 4, "4,5,6,7", 20),
    ("DỄ",         3, "1,2,3",   15),
]
for i, (name, sl, cu_th, std) in enumerate(cd_data):
    r = 39 + i
    put(r, 13, name,  font=NORM, align=LEFT,   border=BOX)
    put(r, 14, sl,    font=NORM, align=CENTER, border=BOX)
    put(r, 15, cu_th, font=NORM, align=LEFT,   border=BOX)
    put(r, 16, std,   font=NORM, align=CENTER, border=BOX)
    put(r, 17, "",    font=NORM, align=CENTER, border=BOX)
    put(r, 18, "",    font=NORM, align=CENTER, border=BOX)

put(42, 13, "TỔNG", font=HEAD, fill=GRY, align=CENTER, border=BOX, span=(1, 3))
put(42, 16, 60,     font=HEAD, fill=GRY, align=CENTER, border=BOX)
put(42, 17, "",     font=HEAD, fill=GRY, align=CENTER, border=BOX)
put(42, 18, "",     font=HEAD, fill=GRY, align=CENTER, border=BOX)

# ── R55+: HƯỚNG DẪN SỬ DỤNG ────────────────────────────────────────────────
guide_rows = [
    "── HƯỚNG DẪN SỬ DỤNG TEMPLATE ──",
    "",
    "1. Điền giá trị vào các ô màu trắng (các ô bôi xám là tiêu đề, không sửa).",
    "2. Phần 'KẾT QUẢ ĐÁNH GIÁ THEO CÂU HỎI RIÊNG BIỆT' (20 câu trắc nghiệm):",
    "   • ĐƠN VỊ: A/B/C/D/E — gán mỗi câu vào 1 trong 5 đơn vị KTCB",
    "   • ĐÁP ÁN: số đáp án đúng (0-5)",
    "   • KẾT QUẢ: O = đúng, X = sai",
    "3. Phần 'ĐÁNH GIÁ THEO ĐƠN VỊ': nhập 'ĐIỂM THỰC TẾ' đạt được cho 5 đơn vị + bài toán có/không lời.",
    "4. Phần 'KẾT QUẢ ĐÁNH GIÁ THEO CÂU HỎI RIÊNG BIỆT' (10 câu tư duy):",
    "   • TỔNG: điểm đạt được cho mỗi câu (so với ĐIỂM CHUẨN của câu đó)",
    "5. Phần 'ĐÁNH GIÁ THEO LĨNH VỰC': nhập 'ĐIỂM THỰC TẾ' đạt được cho 6 năng lực.",
    "6. Phần 'SO SÁNH NĂNG LỰC THEO CẤP ĐỘ': nhập 'ĐIỂM THỰC TẾ' đạt được cho 3 cấp độ (Dễ/Trung bình/Khó).",
    "",
    "── LƯU FILE & IMPORT ──",
    "• File này lưu định dạng .xlsx (Excel). Báo cáo HTML sẽ tự parse khi bạn kéo-thả vào trang.",
    "• Có thể Save As → CSV UTF-8 nếu cần import nhanh hơn (kéo-thả CSV không cần JS XLSX library).",
    "• Tỷ lệ % được báo cáo tự tính, không cần điền vào template.",
]
for i, line in enumerate(guide_rows):
    r = 50 + i
    if line.startswith("──"):
        put(r, 1, line, font=H2, fill=YEL, align=LEFT, span=(1, 10))
    elif line == "":
        continue
    else:
        put(r, 1, line, font=HINT, align=LEFT, span=(1, 10))

# ── freeze top rows ────────────────────────────────────────────────────────
ws.freeze_panes = "A8"

wb.save(OUT)
print(f"[OK] Generated: {OUT}")
print(f"     Open in Excel, fill values, then drag-drop into baocao HTML.")
